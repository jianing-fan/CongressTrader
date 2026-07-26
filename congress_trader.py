"""
Congressional Stock Trade Scraper — Full Buys & Sales (2025+)

Sources:
  1. House Clerk PTR filings   (disclosures-clerk.house.gov)  ← always available
  2. Senate eFTS PTR filings   (efts.senate.gov)              ← best-effort
  3. Capitol Trades public API (api.capitoltrades.com)        ← best-effort

Output: dated Google Sheet with five tabs:
  • All Trades       — every buy and sell, with sector/industry
  • All Purchases    — buys only
  • All Sales        — sales only
  • Top 50 Buys      — top 50 tickers bought & held 30+ days
  • Summary          — per-ticker counts, members, total estimated value
"""

import io
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta
from threading import Lock

import gspread
import yfinance as yf
from google.oauth2.credentials import Credentials as OAuthCredentials
from google.oauth2 import service_account
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build as google_build
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

# ── Config ────────────────────────────────────────────────────────────────────

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
CACHE_DIR   = os.path.join(SCRIPT_DIR, "_pdf_cache")
os.makedirs(CACHE_DIR, exist_ok=True)

HOUSE_BASE  = "https://disclosures-clerk.house.gov"
SENATE_BASE = "https://efts.senate.gov"
CT_API      = "https://api.capitoltrades.com"

START_DATE       = datetime(2025, 1, 1)
NEW_ON_WEEK_START = datetime(2026, 3, 31)
OAUTH_CREDENTIALS_FILE = os.path.join(SCRIPT_DIR, "congress-trades-oauth.json")
OAUTH_TOKEN_FILE       = os.path.join(SCRIPT_DIR, "congress-trades-token.json")
MAX_WORKERS = 12
YAHOO_DELAY = 0.15
ASSET_NAME_REVIEW_LIMIT = 50

# Tickers can change after a disclosure is filed. Keep the disclosed ticker in
# report rows, but use the current Yahoo symbol for metadata lookup.
TICKER_ALIASES = {
    "FI": "FISV",     # Fiserv changed from FI back to FISV in November 2025
    "BRK.B": "BRK-B", # Yahoo uses hyphenated Berkshire class B ticker
}

# Yahoo sometimes recognizes a current symbol but omits sector metadata.
TICKER_INFO_OVERRIDES = {
    "FI": {"sector": "Technology", "industry": "Information Technology Services"},
    "FISV": {"sector": "Technology", "industry": "Information Technology Services"},
}

OWNER_LABELS = {
    "SP": "Spouse", "ME": "Member", "JT": "Joint",
    "DC": "Dependent Child", "OP": "Other",
}

SKIP_ASSET_TYPES = {"GS", "MF", "OL", "HN", "PS", "RE", "FU", "OP"}

AMOUNT_MAP = {
    "$1,001 - $15,000":        (1_001,     15_000,      8_000),
    "$15,001 - $50,000":       (15_001,    50_000,     32_500),
    "$50,001 - $100,000":      (50_001,   100_000,     75_000),
    "$100,001 - $250,000":     (100_001,  250_000,    175_000),
    "$250,001 - $500,000":     (250_001,  500_000,    375_000),
    "$500,001 - $1,000,000":   (500_001, 1_000_000,   750_000),
    "$1,000,001 - $5,000,000": (1_000_001, 5_000_000, 3_000_000),
    "Over $5,000,000":         (5_000_001, 50_000_000, 10_000_000),
}

print_lock    = Lock()
progress_lock = Lock()
_progress     = {"done": 0, "total": 0}
_run_log      = []

# ── Excel style constants ─────────────────────────────────────────────────────

HDR_BUY  = PatternFill("solid", fgColor="1F4E79")   # dark blue  — buys sheet
HDR_SELL = PatternFill("solid", fgColor="7B2D00")   # dark red   — sales
HDR_ALL  = PatternFill("solid", fgColor="203864")   # navy       — all trades
HDR_SUM  = PatternFill("solid", fgColor="375623")   # dark green — summary
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)

ROW_BUY_EVEN  = PatternFill("solid", fgColor="DEEAF1")
ROW_BUY_ODD   = PatternFill("solid", fgColor="FFFFFF")
ROW_SELL_EVEN = PatternFill("solid", fgColor="FCE4D6")
ROW_SELL_ODD  = PatternFill("solid", fgColor="FFFFFF")
ROW_SUM_EVEN  = PatternFill("solid", fgColor="E2EFDA")
ROW_SUM_ODD   = PatternFill("solid", fgColor="FFFFFF")

THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

# ── Logging ───────────────────────────────────────────────────────────────────

def log(msg):
    with print_lock:
        print(msg, flush=True)
        _run_log.append(str(msg))

# ── HTTP helpers ──────────────────────────────────────────────────────────────

_YF_UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
          "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

def fetch(url, data=None, extra_headers=None, timeout=25):
    headers = {"User-Agent": "Mozilla/5.0 (compatible; research-bot/1.0)"}
    if extra_headers:
        headers.update(extra_headers)
    req = urllib.request.Request(url, data=data, headers=headers)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read()


def fetch_pdf(url):
    fname = os.path.join(CACHE_DIR, re.sub(r"[^a-zA-Z0-9._-]", "_", url.split("/")[-1]))
    if os.path.exists(fname):
        with open(fname, "rb") as f:
            return f.read()
    raw = fetch(url)
    with open(fname, "wb") as f:
        f.write(raw)
    return raw

# ── Amount helpers ────────────────────────────────────────────────────────────

def parse_amount(flat_text):
    m = re.search(r"(\$[\d,]+\s*-\s*\$[\d,]+)", flat_text)
    if not m:
        return "Unknown", 0
    raw = re.sub(r"\s+", " ", m.group(1))
    norm = re.sub(r"\s", "", raw)
    for label, (lo, hi, mid) in AMOUNT_MAP.items():
        if re.sub(r"\s", "", label) == norm:
            return label, mid
    nums = [int(x.replace(",", "")) for x in re.findall(r"[\d,]+", raw)]
    mid = (nums[0] + nums[-1]) // 2 if len(nums) >= 2 else 0
    return raw, mid


def clean_asset_name(asset_raw):
    """Clean House PDF asset text and trim bleed-over from prior non-stock rows."""
    # Keep newlines for structure while removing null bytes and other control chars.
    raw = re.sub(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]", "", asset_raw)

    # Page breaks can inject a transaction row plus the table header into the
    # middle of a wrapped asset name, e.g. "Representing one S 10/30/...
    # ID Owner Asset ... common share". Remove that scaffold and keep the
    # asset-name text on both sides.
    raw = re.sub(
        r"(?<=[A-Za-z])\s*[EPS]\s+\d{2}/\d{2}/\d{4}\s*\d{2}/\d{2}/\d{4}\s*"
        r"\$[\d,]+(?:\s*-\s*(?:\$[\d,]+)?)?\s*"
        r"(?:Filing\s+ID\s+#\d+\s*)?"
        r"ID\s+Owner\s+Asset\s+Transaction\s+Type\s+Date\s+Notification\s+Date\s+"
        r"Amount\s+Cap\.\s+Gains\s+>\s*\$200\?\s*",
        " ",
        raw,
        flags=re.DOTALL,
    )

    # Some House PDFs omit the owner code on a stock row that follows an [OT]
    # exchange row. In that case the parser sees one large block, so keep only
    # the trailing asset-name lines before the ticker.
    has_prior_tag = re.search(r"\[[A-Z]{2}\]", raw)
    has_prior_transaction = re.search(r"\b[EPS]\s+\d{2}/\d{2}/\d{4}", raw)
    if has_prior_tag and has_prior_transaction:
        trailing_lines = []
        for line in reversed([ln.strip() for ln in raw.splitlines()]):
            if not line:
                continue
            if trailing_lines and line.endswith("."):
                break
            if re.search(r"\$[\d,]+", line) or re.match(r"^[EPS]\s+\d{2}/\d{2}/\d{4}", line):
                if trailing_lines:
                    break
                continue
            if re.match(r"^(?:F\s*S|S\s*O|D|L)\s*:", line) or line.startswith(("Filing ID", "ID Owner")):
                if trailing_lines:
                    break
                continue
            if trailing_lines and re.search(r"\[[A-Z]{2}\]", line):
                break
            trailing_lines.append(line)
        if trailing_lines:
            raw = "\n".join(reversed(trailing_lines))

    cleaned = re.sub(r"\s+", " ", raw).strip().rstrip(",")
    if len(cleaned) > ASSET_NAME_REVIEW_LIMIT:
        cleaned = re.sub(
            r"\s+(?:F\s*S:\s*New|S\s*O:|D:|L:)\s+.*$",
            "",
            cleaned,
            flags=re.IGNORECASE,
        ).strip().rstrip(",")
        cleaned = re.sub(
            r"\s+[EPS]\s+\d{2}/\d{2}/\d{4}\s*\d{2}/\d{2}/\d{4}\s*"
            r"\$[\d,]+(?:\s*-\s*(?:\$[\d,]+)?)?\s*.*?"
            r"(?:ID Owner Asset Transaction Type Date Notification Date Amount Cap\. Gains > \$200\?|Filing ID #\d+)\s*",
            " ",
            cleaned,
            flags=re.IGNORECASE,
        )
        cleaned = re.sub(r"\s+", " ", cleaned).strip().rstrip(",")
    return cleaned

# ── House Clerk ───────────────────────────────────────────────────────────────

def house_filing_urls(year):
    body = urllib.parse.urlencode({
        "LastName": "", "FilingYear": str(year),
        "State": "", "District": "", "FilingType": "P",
    }).encode()
    html = fetch(
        f"{HOUSE_BASE}/FinancialDisclosure/ViewMemberSearchResult",
        data=body,
        extra_headers={"Content-Type": "application/x-www-form-urlencoded"},
    ).decode(errors="replace")

    filings = []
    for row in re.findall(r'<tr[^>]*role="row"[^>]*>(.*?)</tr>', html, re.DOTALL | re.IGNORECASE):
        link = re.search(r'href="([^"]+\.pdf)"[^>]*>([^<]+)</a>', row, re.IGNORECASE)
        if not link:
            continue
        pdf_path = link.group(1).strip()
        if "/ptr-pdfs/" not in pdf_path:
            continue
        raw = re.sub(r"\s+", " ", link.group(2)).strip()
        raw = re.sub(r"Hon\.\.\s*", "", raw)
        parts = raw.split(",", 1)
        name = f"{parts[1].strip()} {parts[0].strip()}" if len(parts) == 2 else raw
        name = re.sub(r"&[a-z]+;", " ", name).strip()
        filings.append((name, f"{HOUSE_BASE}/{pdf_path.lstrip('/')}"))
    return filings


def parse_pdf(member, pdf_bytes, year, source="House"):
    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        text = "\n".join(p.extract_text() or "" for p in reader.pages)
    except Exception:
        return []

    text = re.sub(r"P\s+T\s+R.*?Amount Cap\.\s*Gains >\s*\$200\?", "", text, flags=re.DOTALL)
    text = re.sub(r"\* For the complete list.*", "", text, flags=re.DOTALL)
    text = re.sub(r"I\s+V\s+D.*", "", text, flags=re.DOTALL)
    text = re.sub(r"F\s+S\s+:.*", "", text)
    text = re.sub(r"S\s+O\s+:.*", "", text)

    trades = []
    for block in re.split(r"\n(?=(?:SP|ME|JT|DC|OP)\s+\S)", text):
        block = block.strip()
        om = re.match(r"^(SP|ME|JT|DC|OP)\s+", block)
        if not om:
            continue
        owner_code = om.group(1)
        rest = block[om.end():]

        at_m = re.search(r"\[(\w+)\]", rest)
        if at_m and at_m.group(1) in SKIP_ASSET_TYPES:
            continue

        # Prefer the ticker directly followed by the asset-type tag [XX] — e.g. (BAH)\n[ST]
        # This avoids picking up a ticker that bled in from the *next* block due to
        # null-byte boilerplate not stripping cleanly in some PDFs.
        ticker_before_tag = re.search(r"\(([A-Z0-9.]{1,6})\)\s*\n?\s*\[", rest)
        if ticker_before_tag:
            ticker = ticker_before_tag.group(1)
        else:
            # Fallback: take the FIRST ticker found (not last, to avoid bleed-over)
            all_tickers_in_block = re.findall(r"\(([A-Z0-9.]{1,6})\)", rest)
            ticker = all_tickers_in_block[0] if all_tickers_in_block else None
        if not ticker:
            continue

        # Asset name: everything before this specific ticker's paren.
        # When the ticker was found via the [XX] tag, use the exact match position
        # to avoid capturing transaction details or description text that follows.
        if ticker_before_tag:
            asset_raw = rest[:ticker_before_tag.start()]
        else:
            asset_raw = rest.split(f"({ticker})")[0]
        asset = clean_asset_name(asset_raw)

        tm = re.search(r"\b(P|S)\s+(\d{2}/\d{2}/\d{4})", block)
        if not tm:
            continue
        tx_type = tm.group(1)
        try:
            tx_date = datetime.strptime(tm.group(2), "%m/%d/%Y")
        except ValueError:
            continue

        amount_str, midpoint = parse_amount(block.replace("\n", " "))

        trades.append({
            "member":     member,
            "chamber":    "House" if source == "House" else "Senate",
            "owner":      OWNER_LABELS.get(owner_code, owner_code),
            "ticker":     ticker,
            "asset":      asset,
            "type":       tx_type,
            "date":       tx_date,
            "amount_str": amount_str,
            "midpoint":   midpoint,
            "source":     source,
            "year":       year,
        })
    return trades

# ── Senate eFTS ───────────────────────────────────────────────────────────────

def senate_filing_urls(year):
    params = urllib.parse.urlencode({
        "q": "", "report_types": "PTR", "filer_type": "senator",
        "date_range": "custom",
        "from_date": f"{year}-01-01", "to_date": f"{year}-12-31",
        "results_count": 250, "start": 0,
    })
    data = json.loads(fetch(f"{SENATE_BASE}/LATEST/search-index?{params}", timeout=15).decode())
    results = []
    for h in data.get("hits", {}).get("hits", []):
        src = h.get("_source", {})
        name = f"{src.get('first_name', '')} {src.get('last_name', '')}".strip()
        pdf_url = src.get("pdf_url", "")
        if name and pdf_url:
            results.append((name, pdf_url))
    return results

# ── Capitol Trades ────────────────────────────────────────────────────────────

def capitol_trades_fetch(days_back=500):
    since = (datetime.today() - timedelta(days=days_back)).strftime("%Y-%m-%d")
    results = []
    for tx_type_filter in ["buy", "sell"]:
        page = 1
        while True:
            params = urllib.parse.urlencode({
                "pageSize": 500, "page": page,
                "txDate_gte": since,
                "txType": tx_type_filter,
                "orderBy": "txDate", "orderDir": "desc",
            })
            try:
                raw = json.loads(fetch(f"{CT_API}/trades?{params}", timeout=20).decode())
            except Exception:
                break
            trades = raw.get("data", [])
            if not trades:
                break
            for t in trades:
                pol   = t.get("politician", {})
                asset = t.get("asset", {})
                ticker = asset.get("assetTicker")
                if not ticker:
                    continue
                ds = t.get("txDate", "")[:10]
                try:
                    tx_date = datetime.strptime(ds, "%Y-%m-%d")
                except ValueError:
                    continue
                if tx_date < START_DATE:
                    continue
                amt = t.get("reportedAmount", {})
                if isinstance(amt, dict):
                    lo, hi = amt.get("lower", 0) or 0, amt.get("upper", 0) or 0
                    mid = (lo + hi) // 2
                    amount_str = f"${lo:,} - ${hi:,}" if lo else "Unknown"
                else:
                    mid, amount_str = 0, "Unknown"
                results.append({
                    "member":     pol.get("name", "Unknown"),
                    "chamber":    pol.get("chamber", "").capitalize() or "Unknown",
                    "owner":      "Member",
                    "ticker":     ticker,
                    "asset":      asset.get("assetName", ticker),
                    "type":       "P" if tx_type_filter == "buy" else "S",
                    "date":       tx_date,
                    "amount_str": amount_str,
                    "midpoint":   mid,
                    "source":     "Capitol Trades",
                    "year":       tx_date.year,
                })
            meta = raw.get("metadata", {})
            total_pages = (meta.get("totalCount", 0) // 500) + 1
            if page >= total_pages or page >= 10:
                break
            page += 1
    return results

# ── Sector + Industry lookup ──────────────────────────────────────────────────

_info_cache = {}   # ticker → {"sector": ..., "industry": ...}

def _etf_sector_from_category(category):
    if not category:
        return "ETF/Fund"
    category_norm = re.sub(r"\s+", " ", str(category)).strip()
    lower = category_norm.lower()
    if lower == "large blend":
        return "Large Blend/Broad Market"

    sector_map = [
        ("technology", "Technology"),
        ("energy", "Energy"),
        ("financial", "Financial Services"),
        ("health", "Healthcare"),
        ("real estate", "Real Estate"),
        ("utilities", "Utilities"),
        ("consumer defensive", "Consumer Defensive"),
        ("consumer cyclical", "Consumer Cyclical"),
        ("communication", "Communication Services"),
        ("industrial", "Industrials"),
        ("materials", "Basic Materials"),
        ("bond", "Bond/Fixed Income"),
        ("government", "Bond/Fixed Income"),
        ("municipal", "Bond/Fixed Income"),
        ("digital assets", "Digital Assets"),
    ]
    for needle, sector in sector_map:
        if needle in lower:
            return sector
    return category_norm

def _looks_like_etf_or_fund(data):
    quote_type = str(data.get("quoteType") or "").upper()
    legal_type = str(data.get("legalType") or "").lower()
    return (
        quote_type in {"ETF", "MUTUALFUND"}
        or "fund" in legal_type
        or bool(data.get("category") or data.get("fundFamily"))
    )

def stock_info(ticker):
    if ticker in _info_cache:
        return _info_cache[ticker]
    lookup_ticker = TICKER_ALIASES.get(ticker, ticker)
    override = TICKER_INFO_OVERRIDES.get(ticker) or TICKER_INFO_OVERRIDES.get(lookup_ticker)
    info = {"sector": "Unknown", "industry": "Unknown"}
    if override:
        info.update(override)
        _info_cache[ticker] = info
        return info
    try:
        data = yf.Ticker(lookup_ticker).info
        sector = data.get("sector") or "Unknown"
        industry = data.get("industry") or "Unknown"
        if (sector == "Unknown" or industry == "Unknown") and _looks_like_etf_or_fund(data):
            category = data.get("category")
            fund_family = data.get("fundFamily")
            etf_sector = _etf_sector_from_category(category)
            sector = etf_sector if sector == "Unknown" else sector
            industry_parts = [str(x).strip() for x in (category, fund_family) if x]
            industry = " / ".join(industry_parts) if industry == "Unknown" and industry_parts else industry
            if industry == "Unknown":
                industry = "ETF/Fund"
        info["sector"] = sector
        info["industry"] = industry
    except Exception:
        pass
    _info_cache[ticker] = info
    time.sleep(YAHOO_DELAY)
    return info

# ── Excel helpers ─────────────────────────────────────────────────────────────

# Regex for characters illegal in Excel cells (control chars except tab/newline)
import re as _re
_ILLEGAL_CHARS = _re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")

def xl_safe(val):
    """Strip null bytes and other Excel-illegal characters from a string value."""
    if isinstance(val, str):
        return _ILLEGAL_CHARS.sub("", val).strip()
    return val


def style_header(ws, row_num, hdr_fill):
    for cell in ws[row_num]:
        cell.font      = HDR_FONT
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN

def style_data_row(ws, row_num, even_fill, odd_fill):
    fill = even_fill if row_num % 2 == 0 else odd_fill
    for cell in ws[row_num]:
        cell.fill      = fill
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border    = THIN

def autofit_columns(ws, min_w=10, max_w=50):
    for col in ws.columns:
        best = min_w
        for cell in col:
            if cell.value:
                best = min(max(best, len(str(cell.value)) + 2), max_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = best

def freeze_and_filter(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

# ── Sheet builders ────────────────────────────────────────────────────────────

ALL_COLS = [
    ("Trade Date",     14),
    ("Member",         28),
    ("Chamber",        10),
    ("Ticker",          8),
    ("Asset Name",     40),
    ("Trade Type",     12),
    ("Amount (Range)", 26),
    ("Midpoint ($)",   16),
    ("Sector",         22),
    ("Industry",       30),
    ("Owner",          16),
    ("Source",         14),
]

def build_all_trades_sheet(ws, trades, title, hdr_fill, even_fill, odd_fill):
    ws.title = title
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28

    for col_idx, (col_name, col_w) in enumerate(ALL_COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w
    style_header(ws, 1, hdr_fill)

    for r, t in enumerate(trades, 2):
        ws.row_dimensions[r].height = 16
        info = _info_cache.get(t["ticker"], {"sector": "Unknown", "industry": "Unknown"})
        row_vals = [
            t["date"].strftime("%Y-%m-%d"),
            t["member"],
            t["chamber"],
            t["ticker"],
            t["asset"],
            "Purchase" if t["type"] == "P" else "Sale",
            t["amount_str"],
            t["midpoint"] if t["midpoint"] else "",
            info["sector"],
            info["industry"],
            t["owner"],
            t["source"],
        ]
        for col_idx, val in enumerate(row_vals, 1):
            ws.cell(row=r, column=col_idx, value=xl_safe(val))
        style_data_row(ws, r, even_fill, odd_fill)

        # Colour the Trade Type cell
        type_cell = ws.cell(row=r, column=6)
        if t["type"] == "P":
            type_cell.font = Font(color="0070C0", bold=True)
        else:
            type_cell.font = Font(color="C00000", bold=True)

    freeze_and_filter(ws)


TOP50_RANK_NOTE = (
    "Rank is by total estimated buy value for the company/ticker, "
    "not by a single trade. Multiple rows can share the same rank when "
    "they are trades for the same ranked company/ticker. This sheet is "
    "sorted by per-trade midpoint amount descending, then by trade date "
    "most recent first, then member."
)

TOP50_COLS = [
    ("Rank",           6),
    ("Ticker",         8),
    ("Company Name",  40),
    ("Trade Date",    14),
    ("Member",        28),
    ("Chamber",       10),
    ("Amount (Range)",26),
    ("Midpoint ($)",  16),
    ("Sector",        22),
    ("Industry",      30),
    ("Source",        14),
]

def build_top50_sheet(ws, rows, ticker_rank):
    ws.title = "Top 50 Buys (Held 30d+)"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28

    for col_idx, (col_name, col_w) in enumerate(TOP50_COLS, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w
    style_header(ws, 1, HDR_BUY)

    current_sector = None
    for r, row in enumerate(rows, 2):
        ws.row_dimensions[r].height = 16
        info = _info_cache.get(row["ticker"], {"sector": "Unknown", "industry": "Unknown"})
        rank = ticker_rank.get(row["ticker"], "")
        row_vals = [
            rank,
            row["ticker"],
            row["date"].strftime("%Y-%m-%d"),
            row["member"],
            row["chamber"],
            row["amount_str"],
            row["midpoint"] if row["midpoint"] else "",
            info["sector"],
            info["industry"],
            row["asset"],
            row["source"],
        ]
        for col_idx, val in enumerate(row_vals, 1):
            ws.cell(row=r, column=col_idx, value=xl_safe(val))

        # Alternate fills, with a slightly stronger tint at sector boundaries
        sector = info["sector"]
        if sector != current_sector:
            current_sector = sector
            fill = PatternFill("solid", fgColor="BDD7EE")  # section header row
        else:
            fill = ROW_BUY_EVEN if r % 2 == 0 else ROW_BUY_ODD
        for cell in ws[r]:
            cell.fill      = fill
            cell.border    = THIN
            cell.alignment = Alignment(vertical="center")

    freeze_and_filter(ws)


SUMMARY_COLS = [
    ("Ticker",          8),
    ("Asset Name",     38),
    ("Sector",         22),
    ("Industry",       30),
    ("# Buys",          8),
    ("# Sales",         8),
    ("Total Buy Est $", 16),
    ("Total Sell Est $",16),
    ("Members (Buyers)",40),
    ("Heaviest Buy Date",14),
]

def build_summary_sheet(ws, all_trades):
    ws.title = "Summary by Ticker"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28

    for col_idx, (col_name, col_w) in enumerate(SUMMARY_COLS, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w
    style_header(ws, 1, HDR_SUM)

    # Aggregate per ticker
    by_ticker = defaultdict(lambda: {
        "asset": "", "buys": 0, "sells": 0,
        "buy_total": 0, "sell_total": 0, "buyers": set()
    })
    for t in all_trades:
        tk = t["ticker"]
        by_ticker[tk]["asset"] = by_ticker[tk]["asset"] or t["asset"]
        if t["type"] == "P":
            by_ticker[tk]["buys"] += 1
            by_ticker[tk]["buy_total"] += t["midpoint"]
            by_ticker[tk]["buyers"].add(t["member"])
        else:
            by_ticker[tk]["sells"] += 1
            by_ticker[tk]["sell_total"] += t["midpoint"]

    # Sort by total buy activity
    sorted_tickers = sorted(
        by_ticker.keys(),
        key=lambda tk: by_ticker[tk]["buy_total"],
        reverse=True,
    )

    for r, tk in enumerate(sorted_tickers, 2):
        ws.row_dimensions[r].height = 16
        d = by_ticker[tk]
        info = _info_cache.get(tk, {"sector": "Unknown", "industry": "Unknown"})
        buyers_str = "; ".join(sorted(d["buyers"]))
        row_vals = [
            tk,
            d["asset"],
            info["sector"],
            info["industry"],
            d["buys"],
            d["sells"],
            d["buy_total"] if d["buy_total"] else "",
            d["sell_total"] if d["sell_total"] else "",
            buyers_str,
        ]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=col_idx, value=xl_safe(val))
            # Format dollar columns
            if col_idx in (7, 8) and isinstance(val, int):
                cell.number_format = '$#,##0'
        style_data_row(ws, r, ROW_SUM_EVEN, ROW_SUM_ODD)

    freeze_and_filter(ws)

# ── Deduplication ─────────────────────────────────────────────────────────────

def deduplicate(trades):
    seen = set()
    out  = []
    for t in trades:
        key = (t["member"], t["ticker"], t["date"].date(), t["type"], t["midpoint"])
        if key not in seen:
            seen.add(key)
            out.append(t)
    return out

# ── Google Sheets output ──────────────────────────────────────────────────────

def _previous_data(gc, drive, folder_id, run_date):
    """Return (carried_rows, prev_tickers) from the most recent previous sheet.

    carried_rows — data rows (no header) from the previous New On Week tab,
                   to be preserved in the current run.
    prev_tickers — set of all tickers that appeared in the previous All Trades tab,
                   used to decide what counts as truly new this run.
    """
    today_title = f"CongressTrader {run_date.strftime('%Y-%m-%d')}"
    results = drive.files().list(
        q=(f"'{folder_id}' in parents and name contains 'CongressTrader'"
           f" and mimeType='application/vnd.google-apps.spreadsheet' and trashed=false"),
        fields="files(id, name)",
        orderBy="name desc",
    ).execute()
    prev_files = [f for f in results.get("files", []) if f["name"] != today_title]
    if not prev_files:
        log("  No previous sheet found — New On Week tab will start fresh")
        return [], set()
    prev = prev_files[0]
    log(f"  Comparing against: {prev['name']}")
    prev_sheet = gc.open_by_key(prev["id"])

    # Tickers from previous All Trades
    prev_tickers = set()
    try:
        ws = prev_sheet.worksheet("All Trades")
        rows = ws.get_all_values()
        if rows:
            ticker_col = rows[0].index("Ticker") if "Ticker" in rows[0] else 3
            prev_tickers = {
                row[ticker_col] for row in rows[1:]
                if len(row) > ticker_col and row[ticker_col]
            }
    except Exception as e:
        log(f"  Warning: could not read previous All Trades: {e}")

    # Rows to carry forward from previous New On Week tab
    carried_rows = []
    try:
        ws = prev_sheet.worksheet("New On Week")
        rows = ws.get_all_values()
        if len(rows) > 1:
            carried_rows = rows[1:]  # drop header row
    except Exception:
        pass  # tab may not exist in older sheets

    return carried_rows, prev_tickers


def _get_credentials():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]

    oauth_token_json = os.environ.get("GOOGLE_OAUTH_TOKEN_JSON")
    if oauth_token_json:
        creds = OAuthCredentials.from_authorized_user_info(json.loads(oauth_token_json), scopes)
        if not creds.valid:
            if creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                raise RuntimeError(
                    "GOOGLE_OAUTH_TOKEN_JSON is present but cannot be refreshed. "
                    "Regenerate the local token and update the GitHub secret."
                )
        return creds

    service_account_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if service_account_json:
        info = json.loads(service_account_json)
        return service_account.Credentials.from_service_account_info(info, scopes=scopes)

    creds = None
    if os.path.exists(OAUTH_TOKEN_FILE):
        creds = OAuthCredentials.from_authorized_user_file(OAUTH_TOKEN_FILE, scopes)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(OAUTH_CREDENTIALS_FILE, scopes)
            creds = flow.run_local_server(port=0)
        with open(OAUTH_TOKEN_FILE, "w") as f:
            f.write(creds.to_json())
    return creds


def write_to_google_sheets(all_sorted, buys_sorted, sells_sorted,
                           top50_rows, ticker_rank, all_trades, run_date):
    creds = _get_credentials()
    gc = gspread.Client(auth=creds)
    drive = google_build("drive", "v3", credentials=creds)

    # Find the CongressTrader folder and the previous sheet's tickers.
    # GitHub Actions can pass the folder ID as a secret to avoid relying on name search.
    folder_id = os.environ.get("GOOGLE_DRIVE_FOLDER_ID")
    if not folder_id:
        folder_results = drive.files().list(
            q="name='CongressTrader' and mimeType='application/vnd.google-apps.folder' and trashed=false",
            fields="files(id)",
        ).execute()
        folder_id = (folder_results.get("files", []) or [{}])[0].get("id")
    carried_rows, prev_tickers = _previous_data(gc, drive, folder_id, run_date) if folder_id else ([], set())

    sheet_title = f"CongressTrader {run_date.strftime('%Y-%m-%d')}"
    log(f"\nCreating Google Sheet: {sheet_title}")
    spreadsheet = gc.create(sheet_title, folder_id=folder_id) if folder_id else gc.create(sheet_title)

    def trade_rows(trades):
        headers = [col for col, _ in ALL_COLS]
        rows = [headers]
        for t in trades:
            info = _info_cache.get(t["ticker"], {"sector": "Unknown", "industry": "Unknown"})
            rows.append([
                t["date"].strftime("%Y-%m-%d"),
                t["member"],
                t["chamber"],
                t["ticker"],
                xl_safe(t["asset"]),
                "Purchase" if t["type"] == "P" else "Sale",
                t["amount_str"],
                t["midpoint"] if t["midpoint"] else "",
                info["sector"],
                info["industry"],
                t["owner"],
                t["source"],
            ])
        return rows

    def top50_data(rows):
        headers = [col for col, _ in TOP50_COLS]
        data = [headers]
        for row in rows:
            info = _info_cache.get(row["ticker"], {"sector": "Unknown", "industry": "Unknown"})
            data.append([
                ticker_rank.get(row["ticker"], ""),
                row["ticker"],
                xl_safe(row["asset"]),
                row["date"].strftime("%Y-%m-%d"),
                row["member"],
                row["chamber"],
                row["amount_str"],
                row["midpoint"] if row["midpoint"] else "",
                info["sector"],
                info["industry"],
                row["source"],
            ])
        return data

    def summary_data(trades):
        headers = [col for col, _ in SUMMARY_COLS]
        by_ticker = defaultdict(lambda: {
            "asset": "", "buys": 0, "sells": 0,
            "buy_total": 0, "sell_total": 0, "buyers": set(),
            "heaviest_buy": 0, "heaviest_buy_date": None,
        })
        for t in trades:
            tk = t["ticker"]
            by_ticker[tk]["asset"] = by_ticker[tk]["asset"] or t["asset"]
            if t["type"] == "P":
                by_ticker[tk]["buys"] += 1
                by_ticker[tk]["buy_total"] += t["midpoint"]
                by_ticker[tk]["buyers"].add(t["member"])
                if t["midpoint"] > by_ticker[tk]["heaviest_buy"]:
                    by_ticker[tk]["heaviest_buy"] = t["midpoint"]
                    by_ticker[tk]["heaviest_buy_date"] = t["date"]
            else:
                by_ticker[tk]["sells"] += 1
                by_ticker[tk]["sell_total"] += t["midpoint"]
        sorted_tickers = sorted(by_ticker, key=lambda tk: by_ticker[tk]["buy_total"], reverse=True)
        data = [headers]
        for tk in sorted_tickers:
            d = by_ticker[tk]
            info = _info_cache.get(tk, {"sector": "Unknown", "industry": "Unknown"})
            heaviest_date = d["heaviest_buy_date"].strftime("%Y-%m-%d") if d["heaviest_buy_date"] else ""
            data.append([
                tk,
                xl_safe(d["asset"]),
                info["sector"],
                info["industry"],
                d["buys"],
                d["sells"],
                d["buy_total"] if d["buy_total"] else "",
                d["sell_total"] if d["sell_total"] else "",
                "; ".join(sorted(d["buyers"])),
                heaviest_date,
            ])
        return data

    def new_on_week_data(buys):
        headers = ["Ticker", "Company Name", "First Purchased", "Trade Volume", "Congress Member"]
        data = [headers]
        # Carry forward rows from previous New On Week tab (only those on/after start date)
        filtered_carried = [
            list(row) for row in carried_rows
            if row and len(row) > 2 and row[2] >= NEW_ON_WEEK_START.strftime("%Y-%m-%d")
        ]
        current_asset_by_ticker = {
            t["ticker"]: xl_safe(t["asset"])
            for t in buys
            if t["date"] >= NEW_ON_WEEK_START and t.get("asset")
        }
        for row in filtered_carried:
            if len(row) > 1 and row[0] in current_asset_by_ticker:
                if "[OT]" in row[1] or len(row[1]) > 120:
                    row[1] = current_asset_by_ticker[row[0]]
        data.extend(filtered_carried)
        # Tickers already present — don't add duplicates
        seen = {row[0] for row in filtered_carried if row}
        # Add purchases that are new relative to the previous run and on/after start date
        for t in buys:
            if (t["date"] >= NEW_ON_WEEK_START
                    and t["ticker"] not in prev_tickers
                    and t["ticker"] not in seen):
                seen.add(t["ticker"])
                data.append([
                    t["ticker"],
                    xl_safe(t["asset"]),
                    t["date"].strftime("%Y-%m-%d"),
                    t["amount_str"],
                    t["member"],
                ])
        return data

    tabs = [
        ("All Trades",              trade_rows(all_sorted)),
        ("All Purchases",           trade_rows(buys_sorted)),
        ("All Sales",               trade_rows(sells_sorted)),
        ("Top 50 Buys (Held 30d+)", top50_data(top50_rows)),
        ("Summary by Ticker",       summary_data(all_trades)),
        ("New On Week",             new_on_week_data(buys_sorted)),
    ]

    # Write first tab into the default sheet, then create the rest
    first_ws = spreadsheet.sheet1
    first_ws.update_title(tabs[0][0])
    first_ws.update(tabs[0][1])
    worksheets = {tabs[0][0]: first_ws}

    def add_cell_note(ws, row_idx, col_idx, note):
        spreadsheet.batch_update({
            "requests": [{
                "updateCells": {
                    "range": {
                        "sheetId": ws.id,
                        "startRowIndex": row_idx,
                        "endRowIndex": row_idx + 1,
                        "startColumnIndex": col_idx,
                        "endColumnIndex": col_idx + 1,
                    },
                    "rows": [{"values": [{"note": note}]}],
                    "fields": "note",
                }
            }]
        })

    def format_google_report():
        header_color = {"red": 0.12, "green": 0.23, "blue": 0.40}
        stripe_color = {"red": 0.93, "green": 0.96, "blue": 1.0}
        white = {"red": 1.0, "green": 1.0, "blue": 1.0}
        buy_blue = {"red": 0.0, "green": 0.28, "blue": 0.58}
        sale_red = {"red": 0.70, "green": 0.12, "blue": 0.12}
        requests = []
        for tab_name, rows in tabs:
            ws = worksheets[tab_name]
            row_count = max(len(rows), 1)
            col_count = max(len(rows[0]) if rows else 1, 1)
            sheet_range = {
                "sheetId": ws.id,
                "startRowIndex": 0,
                "endRowIndex": row_count,
                "startColumnIndex": 0,
                "endColumnIndex": col_count,
            }
            requests.extend([
                {
                    "updateSheetProperties": {
                        "properties": {
                            "sheetId": ws.id,
                            "gridProperties": {"frozenRowCount": 1},
                        },
                        "fields": "gridProperties.frozenRowCount",
                    }
                },
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": ws.id,
                            "startRowIndex": 0,
                            "endRowIndex": 1,
                            "startColumnIndex": 0,
                            "endColumnIndex": col_count,
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": header_color,
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "wrapStrategy": "WRAP",
                                "textFormat": {
                                    "bold": True,
                                    "foregroundColor": white,
                                },
                            }
                        },
                        "fields": (
                            "userEnteredFormat(backgroundColor,horizontalAlignment,"
                            "verticalAlignment,wrapStrategy,textFormat)"
                        ),
                    }
                },
                {"setBasicFilter": {"filter": {"range": sheet_range}}},
                {
                    "addBanding": {
                        "bandedRange": {
                            "range": sheet_range,
                            "rowProperties": {
                                "headerColor": header_color,
                                "firstBandColor": white,
                                "secondBandColor": stripe_color,
                            },
                        }
                    }
                },
                {
                    "autoResizeDimensions": {
                        "dimensions": {
                            "sheetId": ws.id,
                            "dimension": "COLUMNS",
                            "startIndex": 0,
                            "endIndex": col_count,
                        }
                    }
                },
            ])
            if "Trade Type" in rows[0]:
                type_col = rows[0].index("Trade Type")
                type_range = {
                    "sheetId": ws.id,
                    "startRowIndex": 1,
                    "endRowIndex": row_count,
                    "startColumnIndex": type_col,
                    "endColumnIndex": type_col + 1,
                }
                requests.extend([
                    {
                        "addConditionalFormatRule": {
                            "rule": {
                                "ranges": [type_range],
                                "booleanRule": {
                                    "condition": {
                                        "type": "TEXT_EQ",
                                        "values": [{"userEnteredValue": "Purchase"}],
                                    },
                                    "format": {
                                        "textFormat": {
                                            "foregroundColor": buy_blue,
                                            "bold": True,
                                        }
                                    },
                                },
                            },
                            "index": 0,
                        }
                    },
                    {
                        "addConditionalFormatRule": {
                            "rule": {
                                "ranges": [type_range],
                                "booleanRule": {
                                    "condition": {
                                        "type": "TEXT_EQ",
                                        "values": [{"userEnteredValue": "Sale"}],
                                    },
                                    "format": {
                                        "textFormat": {
                                            "foregroundColor": sale_red,
                                            "bold": True,
                                        }
                                    },
                                },
                            },
                            "index": 0,
                        }
                    },
                ])
        spreadsheet.batch_update({"requests": requests})

    for tab_name, rows in tabs[1:]:
        ws = spreadsheet.add_worksheet(title=tab_name, rows=max(len(rows) + 10, 100), cols=20)
        ws.update(rows)
        worksheets[tab_name] = ws
        if tab_name == "Top 50 Buys (Held 30d+)":
            add_cell_note(ws, 0, 0, TOP50_RANK_NOTE)

    format_google_report()

    if folder_id:
        log(f"  Created in Google Drive folder: CongressTrader")
    else:
        log(f"  Warning: folder 'CongressTrader' not found — sheet saved to Drive root")

    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet.id}"
    log(f"\n✓ Google Sheet created: {url}")

    # ── Upload run log to Google Drive ────────────────────────────────────────
    if folder_id:
        log_title = "CongressTrader Run Log"
        log_results = drive.files().list(
            q=(f"'{folder_id}' in parents and name='{log_title}'"
               f" and mimeType='application/vnd.google-apps.spreadsheet' and trashed=false"),
            fields="files(id)",
        ).execute()
        log_files = log_results.get("files", [])
        if log_files:
            log_sheet = gc.open_by_key(log_files[0]["id"])
            log_ws = log_sheet.sheet1
        else:
            log_sheet = gc.create(log_title, folder_id=folder_id)
            log_ws = log_sheet.sheet1
            log_ws.update_title("Run Log")
        rows = [[f"=== Run {run_date.strftime('%Y-%m-%d %H:%M')} ==="]] + \
               [[line] for line in _run_log] + [[""]]
        log_ws.append_rows(rows)

    return url


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    run_date = datetime.today()

    all_trades = []
    years = list(range(START_DATE.year, run_date.year + 1))
    log(f"Collecting trades from {START_DATE.year} through {run_date.year}...")

    # ── 1. House Clerk ────────────────────────────────────────────────────────
    for year in years:
        log(f"\n[House] Fetching PTR index for {year}...")
        try:
            filings = house_filing_urls(year)
        except Exception as e:
            log(f"[House] Failed for {year}: {e}")
            continue
        log(f"[House] {len(filings)} PTR filings for {year}. Downloading PDFs...")

        with progress_lock:
            _progress["done"] = 0
            _progress["total"] = len(filings)

        def _download(args):
            member, url, yr = args
            try:
                trades = parse_pdf(member, fetch_pdf(url), yr, "House")
            except Exception:
                trades = []
            with progress_lock:
                _progress["done"] += 1
                done = _progress["done"]
            return trades

        tasks = [(m, u, year) for m, u in filings]
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
            for batch in pool.map(_download, tasks):
                all_trades.extend(batch)
        log(f"[House {year}] Done — {sum(1 for t in all_trades if t['year']==year and t['source']=='House')} trades")

    # ── 2. Senate eFTS ────────────────────────────────────────────────────────
    log("\n[Senate] Fetching PTR filings...")
    for year in years:
        try:
            filings = senate_filing_urls(year)
            log(f"[Senate] {len(filings)} PTR filings for {year}. Downloading PDFs...")
            tasks = [(m, u, year) for m, u in filings]
            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
                for batch in pool.map(
                    lambda a: parse_pdf(a[0], fetch_pdf(a[1]), a[2], "Senate"), tasks
                ):
                    all_trades.extend(batch)
            log(f"[Senate {year}] Done.")
        except Exception as e:
            log(f"[Senate {year}] Unavailable: {e}")

    # ── 3. Capitol Trades ─────────────────────────────────────────────────────
    log(f"\n[Capitol Trades] Fetching buy & sell trades ({START_DATE.year}+)...")
    try:
        ct = capitol_trades_fetch(days_back=(run_date - START_DATE).days + 31)
        all_trades.extend(ct)
        log(f"[Capitol Trades] {len(ct)} trades fetched.")
    except Exception as e:
        log(f"[Capitol Trades] Unavailable: {e}")

    # ── 4. Deduplicate & filter to 2025+ ───────────────────────────────────
    all_trades = deduplicate(all_trades)
    all_trades = [t for t in all_trades if t["date"] >= START_DATE and t["ticker"]]
    buys  = [t for t in all_trades if t["type"] == "P" and t["midpoint"] > 0]
    sells = [t for t in all_trades if t["type"] == "S"]
    log(f"\nDeduped totals → {len(buys)} purchases · {len(sells)} sales")

    # ── 5. Top 50: buys held 30+ days ────────────────────────────────────────
    sales_index = defaultdict(list)
    for t in sells:
        sales_index[(t["member"], t["ticker"])].append(t["date"])

    held = [
        p for p in buys
        if not any(
            p["date"] <= sd <= p["date"] + timedelta(days=30)
            for sd in sales_index.get((p["member"], p["ticker"]), [])
        )
    ]

    ticker_total = defaultdict(int)
    for p in held:
        ticker_total[p["ticker"]] += p["midpoint"]
    top50_tickers = sorted(ticker_total, key=ticker_total.__getitem__, reverse=True)[:50]

    # ── 6. Sector + industry lookup for ALL unique tickers ───────────────────
    all_tickers = sorted({t["ticker"] for t in all_trades})
    log(f"\nLooking up sector/industry for {len(all_tickers)} unique tickers...")
    for tk in all_tickers:
        stock_info(tk)

    # ── 7. Build top-50 rows (sorted midpoint desc → recent date → member) ───
    ticker_rank = {tk: i+1 for i, tk in enumerate(top50_tickers)}
    top50_rows = []
    for tk in top50_tickers:
        for p in held:
            if p["ticker"] == tk:
                top50_rows.append(p)
    top50_rows.sort(key=lambda r: (-r["midpoint"], -r["date"].timestamp(), r["member"]))

    # ── 8. Build All Trades list (sorted by date desc) ────────────────────────
    all_sorted = sorted(all_trades, key=lambda t: t["date"], reverse=True)
    buys_sorted  = sorted(buys,  key=lambda t: t["date"], reverse=True)
    sells_sorted = sorted(sells, key=lambda t: t["date"], reverse=True)

    # ── 9. Write to Google Sheets ─────────────────────────────────────────────
    url = write_to_google_sheets(
        all_sorted, buys_sorted, sells_sorted,
        top50_rows, ticker_rank, all_trades, run_date,
    )
    log(f"  Sheets: All Trades · All Purchases · All Sales · Top 50 Buys · Summary by Ticker")
    log(f"  Rows:   {len(all_sorted)} total · {len(buys_sorted)} buys · {len(sells_sorted)} sales")
    log(f"  Tickers: {len(all_tickers)} unique")

    # ── 10. Clean up PDF cache ────────────────────────────────────────────────
    import shutil
    if os.path.exists(CACHE_DIR):
        shutil.rmtree(CACHE_DIR)
        log(f"\n✓ PDF cache deleted: {CACHE_DIR}")

    return url


if __name__ == "__main__":
    main()
